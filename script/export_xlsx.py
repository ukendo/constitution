"""Export vymohy/*.md into an Excel workbook under output/."""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

root = Path(__file__).resolve().parent.parent
vymohy = root / "vymohy"
out_dir = root / "output"

readme = (root / "README.md").read_text(encoding="utf-8")
shorts = {}
for m in re.finditer(r"\| \[В-(\d+)\]\([^)]+\) \| (.+?) \|", readme):
    shorts[int(m.group(1))] = m.group(2).strip()


def strip_md_links(text: str) -> str:
    return re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)


def clean_body(raw: str) -> str:
    lines = raw.splitlines()
    i = 0
    if lines and lines[0].startswith("#"):
        i = 1
    while i < len(lines):
        line = lines[i]
        if (
            not line.strip()
            or line.strip() == "---"
            or "← до списку" in line
            or "наступна:" in line
        ):
            i += 1
            continue
        break
    body = "\n".join(lines[i:]).strip()
    body = re.sub(r"\n---\n+", "\n\n", body)
    return strip_md_links(body)


rows = []
for f in sorted(vymohy.glob("V-*.md")):
    m = re.match(r"V-(\d+)\.md", f.name)
    if not m:
        continue
    num = int(m.group(1))
    text = f.read_text(encoding="utf-8")
    title_m = re.match(r"#\s*(.+)", text)
    title = title_m.group(1).strip() if title_m else f"В-{num}"
    rows.append(
        {
            "num": num,
            "id": f"В-{num}",
            "title": title,
            "short": shorts.get(num, ""),
            "body": clean_body(text),
            "file": f"vymohy/{f.name}",
        }
    )

wb = Workbook()

thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
cell_font = Font(name="Calibri", size=11)
wrap = Alignment(wrap_text=True, vertical="top")
header_align = Alignment(wrap_text=True, vertical="center")
title_font = Font(bold=True, name="Calibri", size=14, color="1F4E79")
note_font = Font(name="Calibri", size=10, italic=True, color="666666")


def style_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin


def style_row(ws, row, vals):
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row, col, v)
        cell.font = cell_font
        cell.alignment = wrap
        cell.border = thin


# --- Sheet 1: Список ---
ws1 = wb.active
ws1.title = "Список"
ws1["A1"] = "Вимоги до статуту нової федерації"
ws1["A1"].font = title_font
ws1["A2"] = (
    "Версія 0.1 · зріз з репозиторію (файли vymohy/V-NN.md). "
    "Пріоритет при розбіжностях — окремі markdown-файли."
)
ws1["A2"].font = note_font
ws1.merge_cells("A1:C1")
ws1.merge_cells("A2:C2")

style_header(ws1, 4, ["№", "Короткий опис", "Файл"])
for i, r in enumerate(rows, 5):
    style_row(ws1, i, [r["id"], r["short"], r["file"]])

ws1.column_dimensions["A"].width = 8
ws1.column_dimensions["B"].width = 90
ws1.column_dimensions["C"].width = 18
ws1.row_dimensions[4].height = 22
ws1.freeze_panes = "A5"
ws1.auto_filter.ref = f"A4:C{4 + len(rows)}"

# --- Sheet 2: Повний текст ---
ws2 = wb.create_sheet("Повний текст")
ws2["A1"] = "Повний текст вимог (без навігаційних посилань)"
ws2["A1"].font = title_font
ws2.merge_cells("A1:D1")
ws2["A2"] = (
    "Markdown-посилання замінені на текст мітки. "
    "Таблиці й списки збережені як текст."
)
ws2["A2"].font = note_font
ws2.merge_cells("A2:D2")

style_header(ws2, 4, ["№", "Назва", "Короткий опис", "Повний текст"])
for i, r in enumerate(rows, 5):
    style_row(ws2, i, [r["id"], r["title"], r["short"], r["body"]])
    lines = r["body"].count("\n") + 1
    ws2.row_dimensions[i].height = min(300, max(45, lines * 12))

ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 45
ws2.column_dimensions["C"].width = 45
ws2.column_dimensions["D"].width = 100
ws2.row_dimensions[4].height = 22
ws2.freeze_panes = "A5"
ws2.auto_filter.ref = f"A4:D{4 + len(rows)}"

# --- Sheet 3: Робоча ---
ws3 = wb.create_sheet("Робоча")
ws3["A1"] = "Робочий аркуш для коментарів / пропозицій"
ws3["A1"].font = title_font
ws3.merge_cells("A1:F1")
ws3["A2"] = (
    "Зміни з цього аркуша варто переносити назад у vymohy/V-NN.md через PR — "
    "цей файл не є джерелом істини."
)
ws3["A2"].font = note_font
ws3.merge_cells("A2:F2")

style_header(
    ws3,
    4,
    ["№", "Назва", "Короткий опис", "Статус", "Коментар / пропозиція", "Автор"],
)
for i, r in enumerate(rows, 5):
    style_row(ws3, i, [r["id"], r["title"], r["short"], "", "", ""])

ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 40
ws3.column_dimensions["C"].width = 50
ws3.column_dimensions["D"].width = 16
ws3.column_dimensions["E"].width = 50
ws3.column_dimensions["F"].width = 18
ws3.row_dimensions[4].height = 22
ws3.freeze_panes = "A5"
ws3.auto_filter.ref = f"A4:F{4 + len(rows)}"

dv = DataValidation(
    type="list",
    formula1='"ok,обговорення,редакція,юридичне,відхилено"',
    allow_blank=True,
)
ws3.add_data_validation(dv)
dv.add(f"D5:D{4 + len(rows)}")

out_dir.mkdir(parents=True, exist_ok=True)
out = out_dir / "Вимоги_до_статуту_нової_федерації_v0.1.xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"Rows: {len(rows)}")
for r in rows:
    print(f"  {r['id']}: {r['title'][:70]}")
